import os
import yaml
import re
from datetime import datetime
from typing import Dict, Any, Optional, List
from openpyxl import load_workbook

class ExcelPredictionManager:
    def __init__(self):
        self.predictions_file = "excel_predictions.yaml"
        self.predictions = {}  # {key: {numero, date_heure, victoire, launched, message_id, channel_id}}
        self.last_launched_numero = None  # Dernier numéro lancé pour éviter les consécutifs
        self.load_predictions()

    def import_excel(self, file_path: str) -> Dict[str, Any]:
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active

            imported_count = 0
            skipped_count = 0
            consecutive_skipped = 0
            predictions = {}
            last_numero = None

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1] or not row[2]:
                    continue

                date_heure = row[0]
                numero = row[1]
                victoire = row[2]

                if isinstance(date_heure, datetime):
                    date_str = date_heure.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    date_str = str(date_heure)

                numero_int = int(numero)
                victoire_type = str(victoire).strip()

                prediction_key = f"{numero_int}"

                # Vérifier si déjà lancé
                if prediction_key in self.predictions and self.predictions[prediction_key].get("launched"):
                    skipped_count += 1
                    continue

                # FILTRE CONSÉCUTIFS: Vérifier si numéro actuel = précédent + 1
                # Ex: Si on a 56, on ignore 57, mais on garde 59
                if last_numero is not None and numero_int == last_numero + 1:
                    consecutive_skipped += 1
                    print(f"⚠️ Numéro {numero_int} IGNORÉ À L'IMPORT (consécutif à {last_numero})")
                    # NE PAS mémoriser ce numéro comme last_numero
                    # On continue avec l'ancien last_numero pour détecter le prochain consécutif
                    continue

                predictions[prediction_key] = {
                    "numero": numero_int,
                    "date_heure": date_str,
                    "victoire": victoire_type,
                    "launched": False,
                    "message_id": None,
                    "chat_id": None,
                    "imported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                imported_count += 1
                last_numero = numero_int  # Mémoriser UNIQUEMENT les numéros NON consécutifs

            # REMPLACER complètement les anciennes prédictions par les nouvelles
            self.predictions = predictions
            self.save_predictions()
            print(f"🔄 Anciennes prédictions REMPLACÉES par {len(predictions)} nouvelles prédictions")

            return {
                "success": True,
                "imported": imported_count,
                "skipped": skipped_count,
                "consecutive_skipped": consecutive_skipped,
                "total": len(self.predictions),
                "replaced": True  # Indique que les anciennes ont été remplacées
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

    def save_predictions(self):
        try:
            with open(self.predictions_file, "w", encoding="utf-8") as f:
                yaml.dump(self.predictions, f, allow_unicode=True, default_flow_style=False)
            print(f"✅ Prédictions Excel sauvegardées: {len(self.predictions)} entrées")
        except Exception as e:
            print(f"❌ Erreur sauvegarde prédictions: {e}")

    def _save_predictions(self):
        """Alias pour compatibilité avec main.py"""
        self.save_predictions()

    def load_predictions(self):
        try:
            if os.path.exists(self.predictions_file):
                with open(self.predictions_file, "r", encoding="utf-8") as f:
                    self.predictions = yaml.safe_load(f) or {}
                print(f"✅ Prédictions chargées: {len(self.predictions)} entrées")
            else:
                self.predictions = {}
                print("ℹ️ Aucun fichier de prédictions Excel existant")
        except Exception as e:
            print(f"❌ Erreur chargement prédictions: {e}")
            self.predictions = {}

    def find_close_prediction(self, current_number: int, tolerance: int = 4):
        """
        Trouve une prédiction à lancer quand le canal source affiche un numéro proche AVANT le numéro cible.
        Exemple: Excel #881, Canal source #879 → Lance #881 (diff = +2)
        Tolérance: 0 à 4 parties d'écart
        IMPORTANT: Ignore les numéros consécutifs (ex: 56→57 ignoré, on passe directement à 59)
        """
        try:
            closest_pred = None
            min_diff = float('inf')

            for key, pred in self.predictions.items():
                if pred["launched"]:
                    continue

                pred_numero = pred["numero"]
                # Calculer la différence: pred_numero - current_number
                # Si canal=879 et pred=881, diff=+2 (canal est 2 parties AVANT)
                diff = pred_numero - current_number

                # Vérifier si le canal source est entre 0 et 4 parties AVANT le numéro cible
                if 0 <= diff <= tolerance:
                    # FILTRE PRINCIPAL: Vérifier si ce n'est pas un numéro consécutif du dernier prédit
                    if self.last_launched_numero and pred_numero == self.last_launched_numero + 1:
                        print(f"⚠️ Numéro {pred_numero} IGNORÉ AU LANCEMENT (consécutif à {self.last_launched_numero})")
                        # Marquer comme lancé pour éviter de le relancer plus tard
                        pred["launched"] = True
                        pred["skipped_consecutive"] = True
                        self.save_predictions()
                        continue

                    # Garder la prédiction la plus proche (priorité au plus petit écart)
                    if diff < min_diff:
                        min_diff = diff
                        closest_pred = {"key": key, "prediction": pred}
                        print(f"✅ Prédiction trouvée: #{pred_numero} (canal #{current_number}, écart +{diff})")

            return closest_pred
        except Exception as e:
            print(f"Erreur find_close_prediction: {e}")
            return None

    def mark_as_launched(self, key: str, message_id: int, channel_id: int):
        """Marque une prédiction comme lancée"""
        if key in self.predictions:
            self.predictions[key]["launched"] = True
            self.predictions[key]["message_id"] = message_id
            self.predictions[key]["channel_id"] = channel_id
            self.predictions[key]["current_offset"] = 0  # Commence avec offset 0
            self.last_launched_numero = self.predictions[key]["numero"]
            self.save_predictions()

    def _extract_points(self, message_text: str):
        """Extrait les points du joueur et du banquier depuis le message de résultat"""
        try:
            # Format: #N249. ✅8(6♦️2♠️) - 1(5♦️6♦️) ou #N253. 2(2♣️J♥️) - ✅9(3♣️6♦️)
            # Match nul: #N252. 7(3♠️4♣️) 🔰 7(A♦️6♦️)

            # Pattern pour extraire les points
            pattern = r'(\d+)\([^)]+\)'
            matches = re.findall(pattern, message_text)

            if len(matches) >= 2:
                # Le premier groupe (avant le tiret) est TOUJOURS le joueur
                # Le second groupe (après le tiret) est TOUJOURS le banquier
                joueur_point = int(matches[0])
                banquier_point = int(matches[1])

                # Validation STRICTE: vérifier que le ✅ correspond bien au gagnant
                parts = message_text.split('-') if '-' in message_text else message_text.split('🔰')

                if '✅' in message_text and not '🔰' in message_text:
                    # Vérifier la cohérence entre ✅ et les points
                    if '✅' in parts[0]:
                        # ✅ avant le tiret → joueur DOIT avoir gagné
                        if joueur_point <= banquier_point:
                            print(f"❌ Incohérence CRITIQUE: ✅ sur joueur mais points joueur ({joueur_point}) <= banquier ({banquier_point}) - REJET")
                            return None, None
                    elif len(parts) > 1 and '✅' in parts[1]:
                        # ✅ après le tiret → banquier DOIT avoir gagné
                        if banquier_point <= joueur_point:
                            print(f"❌ Incohérence CRITIQUE: ✅ sur banquier mais points banquier ({banquier_point}) <= joueur ({joueur_point}) - REJET")
                            return None, None

                return joueur_point, banquier_point

            return None, None
        except Exception as e:
            print(f"Erreur extraction points: {e}")
            return None, None

    def verify_excel_prediction(self, game_number: int, message_text: str, predicted_numero: int, expected_winner: str, current_offset: int):
        """
        Vérifie une prédiction Excel avec calcul des points pour déterminer le gagnant.

        Args:
            game_number: Numéro du jeu actuel
            message_text: Texte du message de résultat
            predicted_numero: Numéro prédit
            expected_winner: Gagnant attendu (joueur/banquier)
            current_offset: Offset réel calculé (0, 1, 2)

        Returns:
            tuple: (status, should_continue)
                - status: '✅0️⃣', '✅1️⃣', '✅2️⃣', '❌', ou None
                - should_continue: True si on doit continuer à vérifier, False si terminé
        """
        try:
            # L'offset est maintenant calculé en amont, on vérifie juste qu'il correspond
            real_offset = game_number - predicted_numero

            # Validation de cohérence
            if real_offset != current_offset:
                print(f"⚠️ ERREUR: offset incohérent - real={real_offset}, current={current_offset}")
                return None, True

            # Si le jeu est avant la prédiction, continuer à attendre
            if real_offset < 0:
                print(f"⏭️ Jeu #{game_number} est AVANT la prédiction #{predicted_numero}")
                return None, True

            # Si l'offset est trop grand, échec définitif
            if real_offset > 2:
                print(f"❌ Prédiction Excel #{predicted_numero}: offset {real_offset} > 2, échec définitif")
                return '❌', False

            # **NOUVEAU**: Détecter match nul (🔰) - IGNORER et continuer
            if '🔰' in message_text:
                print(f"⚪ Match NUL détecté sur jeu #{game_number} - IGNORÉ, on continue")
                return None, True

            # ATTENDRE que le message soit finalisé avec ✅
            if '⏰' in message_text:
                print(f"⏰ Message en cours d'édition #{game_number}, attente de finalisation avec ✅")
                return None, True

            # Vérifier si le message contient ✅ (message finalisé)
            if '✅' not in message_text:
                print(f"⚠️ Message #{game_number} sans ✅ (non finalisé), on continue à attendre")
                return None, True

            # Extraire les points
            joueur_point, banquier_point = self._extract_points(message_text)

            if joueur_point is None or banquier_point is None:
                # Vérifier si c'est un message avec ✅ mais sans points clairs
                if '✅' in message_text and '🔰' not in message_text:
                    print(f"❌ CRITIQUE: Message avec ✅ incohérent - échec immédiat")
                    return '❌', True
                else:
                    print(f"⚠️ Impossible d'extraire les points, attente...")
                    return None, True

            # Déterminer le gagnant réel selon les points
            if joueur_point > banquier_point:
                actual_winner = "joueur"
            elif banquier_point > joueur_point:
                actual_winner = "banquier"
            else:
                # Match nul par égalité de points - ignorer aussi
                print(f"⚪ Match NUL détecté (égalité J:{joueur_point} = B:{banquier_point}) - IGNORÉ")
                return None, True

            # Comparer avec le gagnant attendu
            expected = "banquier" if "banquier" in expected_winner.lower() else "joueur"

            print(f"📊 Points: Joueur={joueur_point}, Banquier={banquier_point} → Gagnant: {actual_winner}, Attendu: {expected}")

            if actual_winner != expected:
                print(f"❌ OFFSET {current_offset}: gagnant INCORRECT ({actual_winner} ≠ {expected})")
                return None, False

            # ✅ SUCCÈS ! Gagnant correct sur cet offset
            print(f"✅ OFFSET {current_offset}: gagnant CORRECT ({actual_winner} = {expected}) - ARRÊT")

            if current_offset == 0:
                return '✅0️⃣', True
            elif current_offset == 1:
                return '✅1️⃣', True
            elif current_offset == 2:
                return '✅2️⃣', True
            else:
                # Sécurité (ne devrait pas arriver ici)
                return '❌', True

        except Exception as e:
            print(f"Erreur verify_excel_prediction: {e}")
            return None, False

    def get_prediction_format(self, victoire: str) -> str:
        victoire_lower = victoire.lower()

        if "joueur" in victoire_lower or "player" in victoire_lower:
            return "👗 𝐕𝟏👗"
        elif "banquier" in victoire_lower or "banker" in victoire_lower:
            return "👗 𝐕2👗"
        else:
            return "👗 𝐕𝟏👗"

    def get_pending_predictions(self) -> List[Dict[str, Any]]:
        pending = []
        for key, pred in self.predictions.items():
            if not pred["launched"]:
                pending.append({
                    "key": key,
                    "numero": pred["numero"],
                    "victoire": pred["victoire"],
                    "date_heure": pred["date_heure"]
                })
        return sorted(pending, key=lambda x: x["numero"])

    def get_stats(self) -> Dict[str, int]:
        total = len(self.predictions)
        launched = sum(1 for p in self.predictions.values() if p["launched"])
        pending = total - launched

        return {
            "total": total,
            "launched": launched,
            "pending": pending
        }

    def clear_predictions(self):
        self.predictions = {}
        self.save_predictions()
        print("🗑️ Toutes les prédictions Excel ont été effacées")